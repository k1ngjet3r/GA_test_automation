from openpyxl import load_workbook
from openpyxl import Workbook
from datetime import datetime
from pushbullet import Pushbullet
from datetime import date
from gtts import gTTS
import speech_recognition as sr
import pyttsx3
import time
import os
import re

# mic setup
r = sr.Recognizer()
mic = sr.Microphone(device_index=1)

today = str(date.today())
exe_date = today[:4] + today[5:7] + today[8:]

# Setting up the output file
output_name = 'SMS_auto_result_{}.xlsx'.format(exe_date)
sheet_titles = ['Online_In', 'Offline_In', 'Online_Out', 'Offline_Out']
out = Workbook()
out.active
for name in sheet_titles:
    out.create_sheet(name)
    out[name].append(
        ['TCID', 'Test Step', 'Time of Execution', 'GA_respond', 'Result'])


def ambient_noise(recog, mic):
    with mic as source:
        print('Adjusting amnient noise...')
        # Collect and adjust the ambient nosie threadhole
        recog.adjust_for_ambient_noise(source, duration=5)


def stt(recognizer, microphone):
    with microphone as source:
        # Collecting the respond with 150 seconds of waiting time
        print('Collecting Respond...')
        audio = recognizer.listen(source, timeout=300)

    response = {'success': True, "error": None, "transcription": None}

    try:
        # set the recognize language to English and convert the speech to text
        recog = recognizer.recognize_google(audio, language='en-US')
        response["transcription"] = recog

    except sr.RequestError:
        response['success'] = False
        response['error'] = 'API unavailable'

    except sr.UnknownValueError:
        response['success'] = False
        response['error'] = "Unable to recognitized the speech"

    except:
        response['success'] = False
        response['error'] = 'No respond from Google Assistant'
    return response['transcription']


def tts(step):
    engine = pyttsx3.init()
    engine.setProperty('rate', 105)
    engine.say(step)
    engine.runAndWait()


def filename_formater(date):
    return date[:4] + '_' + date[5:7] + '_' + date[8:10] + '_' + date[11:13] + '_' + date[14:16] + '_' + date[17:19]

# activate Google Assistant using adb command


def activate_ga():
    os.system('adb shell am start -n com.google.android.carassistant/com.google.android.apps.gsa.binaries.auto.app.voiceplate.VoicePlateActivity')


# reset Google Assistant session
def reset():
    os.system('adb shell input keyevent 3')
    time.sleep(3)


def push_noti(message):
    # load the key from the pushbullet_api_key.txt
    key = open('pushbullet_api_key.txt', 'r').read()
    pb = Pushbullet(key)
    dev = pb.get_device('Google Pixel 4a (5G)')
    dev.push_note('Automation Notification', message)


class Automation():
    def __init__(self, input_file, output_file):
        self.input_file = str(input_file)
        self.output_file = str(output_file)
        # loading the test cases file
        self.sheet = load_workbook(self.input_file).active
        # Create an empty workbook for output
        self.wb = Workbook()
        self.wb.active
        # Create the sheetname
        self.wb.create_sheet('AutoResult')
        # Append title of the sheet
        self.wb['AutoResult'].append(
            ['TCID', 'Test Step', 'Time of Execution', 'GA_responds'])

    def step_detail(self, step, msg):
        step = step.lower()

        test_command = []
        default_name = ['walter', 'white']
        default_number = ['0912345678']
        '''
            Command should be something like this
            test_command = ['SMS/send msg', <name>, <type>, 'msg']
        '''
        # determine the stage 1 command
        if re.search('sms', step):
            test_command.append('SMS ')
        elif re.search('call', step) or re.search('dial', step) or re.search('dail', step):
            test_command.append('call')
        else:
            test_command.append('Send message to ')

        # determine the name format
        if re.search('<number>', step) or re.search('<phone_number>', step):
            test_command.append(default_number[0])
        elif re.search('<fn>', step):
            test_command.append(default_name[0])
        elif re.search('<ln>', step):
            test_command.append(default_name[1])
        else:
            test_command.append(default_name[0] + ' ' + default_name[1])

        # determine the phone type
        if re.search('work', step) and re.search('the work', step) is False:
            test_command.append('work')
        
        elif re.search('home', step):
            test_command.append('home')

        else:
            test_command.append('mobile')

        test_command.append(msg)

        return test_command

    # When receiving "Sure, what is the message?"

    def whats_the_msg(self, msg, result):
        print("--> What's the msg?")
        print('   {}'.format(msg))
        tts(msg)
        
        ga_msg_respond = stt(r, mic)
        result.append(ga_msg_respond)
        if re.search('change', ga_msg_respond):
            print('--> ...do you want to send it or change it?')
            print('   send')
            tts('send')
            send_msg_respond = stt(r, mic)
            result.append(send_msg_respond)
            if re.search('sending', send_msg_respond):
                print('--> Sending the message.')
                result.append('Pass')
            else:
                print('Unrecognitized respond: {}'.format(send_msg_respond))
                result.append('Something went wrong')
        else:
            print('Unrecognitized respond: {}'.format(ga_msg_respond))
            result.append('Something went wrong')

    # when receiving "Okay, home, work, or mobile?"
    def phone_type(self, phonetype, msg, result):
        print("--> Okay, home, work, or mobile?")
        print('   {}'.format(phonetype))
        tts(phonetype)
        phone_type_respond = stt(r, mic)
        result.append(phone_type_respond)
        if re.search('what', phone_type_respond):
            self.whats_the_msg(msg, result)

        elif re.search('calling', phone_type_respond):
            self.calling(result)
        
        else:
            print('Unrecognitized respond: {}'.format(phone_type_respond))
            result.append('Something went wrong')

    # when receiving "... for example, the first one"
    def first_one(self, phonetype, msg, result):
        print("--> ...for example, the first one?")
        tts('the first one')
        phone_type_respond = stt(r, mic)
        result.append(phone_type_respond)
        if re.search('what', phone_type_respond):
            self.whats_the_msg(msg, result)
        elif re.search('calling', phone_type_respond):
            self.calling(result)
        else:
            result.append('Something went wrong')

    # when receiving "who do you want to message to?"
    def who(self, name, phonetype, msg, result):
        print("--> Who do you want to message to?")
        print('   {}'.format(name))
        tts(name)
        who_respond = stt(r, mic)
        result.append(who_respond)
        if re.search('is that', who_respond):
            self.is_that(msg, result)

        # Home, work, or mobile?
        elif re.search('or mobile', who_respond):
            self.phone_type(phonetype, msg, result)

        # What's the msg?
        elif re.search('what', who_respond):
            self.whats_the_msg(msg, result)

        elif re.search('calling', who_respond):
            self.calling(result)

        else:
            print('Unrecognitized respond: {}'.format(who_respond))
            result.append('Something went wrong')

    def is_that(self, msg, result):
        print("--> is that <name>'s <type>?")
        print('   {}'.format("yes"))
        tts('yes')
        is_that_respond = stt(r, mic)
        result.append(is_that_respond)
        if re.search('what', is_that_respond):
            self.whats_the_msg(msg, result)
        
        elif re.search('calling', is_that_respond):
            self.calling(result)
        
        else:
            print('Unrecognitized respond: {}'.format(is_that_respond))
            result.append('Something went wrong')

    def calling(self, result):
        print('>>> Call made, ending the call in 5 seconds...')
        # let is ring for 5 second and end the call
        time.sleep(5)
        os.system('adb shell input keyevent 6')
        result.append('Pass')

    def sms_attempt(self, test_command, result):
        # First attempt
        # adjusting the ambinat noise
        ambient_noise(r, mic)
        # Activate Google Assistant
        activate_ga()
        # wait 0.6 second before giving command
        time.sleep(0.8)
        # Give command "SMS/Send msg to <name>"
        tts(test_command[0]+' '+test_command[1])
        # collecting respond from Google Assistant
        time.sleep(1)
        respond = stt(r, mic)
        result.append(respond)

        if re.search('sorry', respond):
            self.who(test_command[1], test_command[2], test_command[3], result)

        elif re.search('or mobile', respond):
            self.phone_type(test_command[2], test_command[3], result)

        elif re.search('is that', respond):
            self.is_that(test_command[3], result)

        elif re.search('what', respond):
            self.whats_the_msg(test_command[3], result)

        elif re.search('calling', respond):
            self.calling(result)

        # first attempt failed, performing second attempt
        else:
            reset()
            # adjusting the ambinat noise
            ambient_noise(r, mic)
            # Activate Google Assistant
            activate_ga()
            # wait 0.6 second before giving command
            time.sleep(0.6)
            # Give command "SMS/Send msg to <name>"
            tts(test_command[0]+' '+test_command[1])
            # collecting respond from Google Assistant
            respond = stt(r, mic)
            result.append(respond)

            if re.search('sorry', respond):
                self.who(test_command[1], test_command[2], test_command[3], result)

            elif re.search('or mobile', respond):
                self.phone_type(test_command[2], test_command[3], result)

            elif re.search('first one', respond):
                self.first_one(test_command[2], test_command[3], result)

            elif re.search('is that', respond):
                self.is_that(test_command[3], result)

            elif re.search('what', respond):
                self.whats_the_msg(test_command[3], result)

            else:
                result.append("Fail")

    def execute(self, sheet_name):
        cases = {str(tcid): str(step) for tcid, step in self.sheet.iter_rows(max_col=2, values_only=True) if tcid is not None}

        current_case = 1
        # Iterate the case and feed it to the main loop
        for tcid in cases:
            ToEx = datetime.now()
            print('{} Execute Case: {}'.format(ToEx, tcid))

            msg = 'Test case nomber {}'.format(str(current_case))

            current_case += 1

            test_command = self.step_detail(cases[tcid], msg)

            result = [tcid, str(test_command), ToEx]

            '''
                Command should be something like this
                test_command = ['SMS/send msg', <name>, <type>, 'msg']
            '''

            try:
                self.sms_attempt(test_command, reset)
            
            except TypeError:
                result.append("something went wrong ")


            out[sheet_name].append(result)


test = Automation('Android_online_in.xlsx', 'output.xlsx')

test.execute('Online_In')