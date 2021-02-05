from openpyxl import load_workbook
from openpyxl import Workbook
from datetime import datetime
from pushbullet import Pushbullet
from datetime import date
from gtts import gTTS
import speech_recognition as sr
import pyttsx3
import time
import os
import re

class Automation():
    def __init__(self, test_case_file, output_file):
        self.test_case_file = load_workbook(test_case_file).active
        self.output_file = Workbook(write_only=True).active



