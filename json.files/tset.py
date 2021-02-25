from openpyxl import load_workbook

wb = load_workbook('W09_sorted.xlsx')

print(str([i[0] for i in wb['Fuel_sim'].iter_rows(max_col=1, max_row=143, values_only=True) if i[0]!=None]).replace("'", '"').replace(' ', '\n'))
