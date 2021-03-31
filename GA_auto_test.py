from openpyxl import load_workbook
from openpyxl import Workbook
from datetime import datetime
from datetime import date
from pushbullet import Pushbullet
from gtts import gTTS
import speech_recognition as sr
import subprocess
import sys
import cv2
import pyttsx3
import time
import os
import re
import sys

r = sr.Recognizer()
mic = sr.Microphone(device_index=1)

today = str(date.today())
exe_date = today[:4] + today[5:7] + today[8:]

path = 'C:\\control\\'

output_name = 'auto_result_{}.xlsx'.format(exe_date)
sheet_titles = ['Online_In', 'Offline_In', 'Online_Out', 'Offline_Out']
out = Workbook()
out.active


def activate_ga():
    os.system('adb shell am start -n com.google.android.carassistant/com.google.android.apps.gsa.binaries.auto.app.voiceplate.VoicePlateActivity')


for name in sheet_titles:
    out.create_sheet(name)
    out[name].append(
        ['TCID', 'Test Step', 'Time of Execution', 'GA_respond', 'Result'])


def ambient_noise(recog, mic):
    with mic as source:
        print('Adjusting amnient noise...')
        # Collect and adjust the ambient nosie threadhole
        recog.adjust_for_ambient_noise(source, duration=5)


def stt(recognizer, microphone):
    with microphone as source:
        # print('Adjusting ambient noise')
        # recognizer.adjust_for_ambient_noise(source, duration = 1)

        # Collecting the respond with 150 seconds of waiting time
        print('    Collecting Respond...')
        audio = recognizer.listen(source, timeout=200)

    response = {'success': True, "error": None, "transcription": None}

    try:
        # set the recognize language to English and convert the speech to text
        recog = recognizer.recognize_google(audio, language='en-UK')
        response["transcription"] = recog

    except sr.RequestError:
        response['success'] = False
        response['error'] = 'API unavailable'

    except sr.UnknownValueError:
        response['success'] = False
        response['error'] = "Unable to recognitized the speech"

    except:
        response['success'] = False
        response['error'] = 'No respond from Google Assistant'
    return response


def tts(step):
    engine = pyttsx3.init()
    engine.setProperty('rate', 105)
    # say "hey google" first and than say the command with 0.5 second delay
    activate_ga()
    time.sleep(1.5)
    # Giving the commend
    engine.say(step)
    engine.runAndWait()


def filename_formater(date):
    y = date[:4]
    m = date[5:7]
    d = date[8:10]
    h = date[11:13]
    minute = date[14:16]
    s = date[17:19]
    return y + '_' + m + '_' + d + '_' + h + '_' + minute + '_' + s


def capturing(tcid):
    cam = cv2.VideoCapture(0)
    cv2.namedWindow("test")
    img_counter = 0

    while True:
        ret, frame = cam.read()

        if not ret:
            print("failed to grab frame")
            break
        cv2.imshow("test", frame)
        cv2.waitKey(1)

        if img_counter < 1:
            img_name = "{}.png".format(tcid)
            cv2.imwrite(img_name, frame)
            print("{} written!".format(img_name))
            img_counter += 1
        else:
            break

    cam.release()
    cv2.destroyAllWindows()


def screenshot(status):
    os.system('$ex = Test-Path C:\Screenshot -PathType Container')
    os.system('if($ex -ne 1) {mkdir C:\Screenshot}')
    os.system('adb shell screencap -p /sdcard/{}.png'.format(status))
    os.system('adb pull /sdcard/{}.png /Screenshot'.format(status))
    os.system('adb shell rm /sdcard/{}.png'.format(status))


def push_noti(message):
    # load the key from the pushbullet_api_key.txt
    key = open('pushbullet_api_key.txt', 'r').read()
    pb = Pushbullet(key)
    dev = pb.get_device('Google Pixel 4a (5G)')
    dev.push_note('Automation Notification', message)


def turn_on_wifi():
    os.system('adb root')
    os.system('adb shell "svc wifi enable"')


def turn_off_wifi():
    os.system('adb root')
    os.system('adb shell "svc wifi disable"')


def sign_out():
    p = subprocess.Popen(
        ['powershell.exe', path+'SignOut.ps1'])
    p.communicate()


def sign_in():
    p = subprocess.Popen(
        ['powershell.exe', path+'SignIn.ps1'])
    p.communicate()


def reset():
    os.system('adb shell input keyevent 3')
    time.sleep(3)


def match_slice(sentence, keywords):
    sen = sentence.lower()
    for key in keywords:
        if re.search(key, sen):
            return True
    return False


def analysis(respond):
    fail_keyword = ["offline", "can't do that", 'sorry',
                    'trouble', 'wrong', 'try again', "don't", "none"]
    if match_slice(respond.lower(), fail_keyword):
        return False
    return True


def system_check():
    tts("hello")

    time.sleep(5)

    turn_off_wifi()

    time.sleep(5)

    turn_on_wifi()

    time.sleep(5)

    sign_out()

    sign_in()

    print('system check done')

class Automation():
    def __init__(self, input_file, sheet_name):
        self.input_file = str(input_file)
        # loading the test cases file
        self.sheet = load_workbook(self.input_file)[sheet_name]

    def execute(self, sheet_name):
        # Generate the dictionary for the case
        cases = {str(tcid): str(step) for tcid, step in self.sheet.iter_rows(
            max_col=2, values_only=True) if tcid is not None}
        case_amount = len(cases)

        print('executing file {}'.format(self.input_file))

        num = 0

        # Iterate the case and feed it to the main loop
        for tcid in cases:
            try:
                num += 1
                ToEx = datetime.now()
                print("Case {}/{}".format(num, case_amount))
                print('{} Execute Case {}'.format(ToEx, tcid))

                # Adjusting the ambient noise threadhole for 5 seconds
                ambient_noise(r, mic)

                result = [tcid, cases[tcid], ToEx]

                # Formatting the command
                text = cases[tcid].replace('\n', ' ')
                text = text.replace('”', '"')
                text = text.replace('“', '"')
                text = text.replace('/', '')
                text = text.split('"')
                if len(text) == 1:
                    text = text[0]
                else:
                    text = text[-2]
                print('    Commend: {}'.format(text))

                # Generate the speech
                tts(text)

                # Reciving Respond
                respond = stt(r, mic)
                print("    Respond: {}".format(str(respond['transcription'])))

                # Capturing the image if the computer captured the respond
                if respond['transcription'] is not None:
                    # capturing(tcid)
                    print('    respond get')

                # If the computer cannot get the respond, it will execute the case again
                else:
                    # Try to perform the test case again
                    print('    ===> Try to perform the case again')
                    
                    # clear the previous condition and recalibrate the ambient noise threadhole
                    reset()
                    ambient_noise(r, mic)

                    tts(text)

                    # Reciving Respond
                    respond = stt(r, mic)
                    print("     Respond: {}".format(
                        str(respond['transcription'])))

                    if respond['transcription'] is not None:
                        # capturing(tcid)
                        print('respond get')

                    # Won't capture photo if the respond is still none
                    else:
                        print('     Fail to perform the test case {}'.format(tcid))
                        error_time = str(datetime.now())[:-7]
                        error_msg = '{}\nFail to perform case {}/{}\nTCID: {}'.format(
                            error_time, num, case_amount, tcid)
                        push_noti(error_msg)

                # Append the result to the output excel file
                result.append(str(respond['transcription']))
                if analysis(str(respond['transcription'])):
                    result.append("Pass")
                out[sheet_name].append(result)
                out.save(output_name)
                print(respond)
                print(
                    '====================================================================================')
                time.sleep(3)

            except:
                print('    Something went wrong, skipping case: {}'.format(tcid))
                push_noti('Error occured when executing case: {}'.format(tcid))


# Create "auto_log.txt" for storing log
# sys.stdout = open('auto_log.txt', 'w')

plan = 'W14_auto.xlsx'

push_noti('Execution Started')
# online_signin
print('Executing Online/Sign In cases')
push_noti('Executing online_in.xlsx')
test_1 = Automation(plan, 'Online_In')
test_1.execute(sheet_titles[0])
push_noti('Stage 1 finished!')

# disconnect WiFi
print('***Disconnecting WiFi***')
turn_off_wifi()
time.sleep(3)
screenshot('offline_in')

# offline_signin
print('Executing Offline/Sign In cases')
push_noti('Executing offline_in.xlsx')
test_2 = Automation(plan, 'Offline_In')
test_2.execute(sheet_titles[1])
push_noti('Stage 2 finished!')

# connect WiFi
print(' ')
print('***Connecting WiFi***')
turn_on_wifi()
time.sleep(10)

# Sign out google account
print('***Signing out google account***')
sign_out()
print(' ')
print(' ')

time.sleep(3)
screenshot('Online_out')

# online_signout
print('Executing Online/Sign out cases')
push_noti('Executing online_out.xlsx')
test_3 = Automation(plan, 'Online_Out')
test_3.execute(sheet_titles[2])
push_noti('Stage 3 finished!')

# disconnect WiFi
print(' ')
print('***Disconnecting WiFi***')
turn_off_wifi()
print(' ')
print(' ')

time.sleep(3)
screenshot('Offline_out')

# offline_signout
print('Executing Offline/Sign out cases')
push_noti('Executing ac_offline_out.xlsx')
test_4 = Automation(plan, 'Offline_Out')
test_4.execute(sheet_titles[3])

push_noti('All test cases executed.')
# Export the result
print('Saving the file {}'.format(output_name))
out.save(output_name)

# sys.stdout.close()
