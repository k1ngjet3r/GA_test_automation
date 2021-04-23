from openpyxl import load_workbook
import json


class Dump:
    def __init__(self, file_name):
        self.wb = load_workbook(file_name)

    def intel(self):
        target_sheet = ['trailer', 'DID', 'User_Build']
        intel = {}
        for sheet in target_sheet:
            print('generate the {} list'.format(sheet))
            temp = []
            for row in self.wb[sheet].iter_rows(max_col=1, values_only=True):
                try:
                    temp.append(row[0].lower())
                except:
                    break
            intel[sheet] = temp[1:]
        return intel

    def dump(self):
        with open('tcid_and_sheet.json', 'w') as outfile:
            json.dump(self.intel(), outfile)


if __name__ == '__main__':
    Dump('W17_Main_sorted.xlsx').dump()
