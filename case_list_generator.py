from openpyxl import load_workbook, Workbook

# Loading the full-list file
full_list = load_workbook('W13_sorted.xlsx')['auto']

# Creating the automation-related cases spreadsheet
wb = Workbook()
sheet_name = ['Online_In', 'Offline_In', 'Online_Out', 'Offline_Out']
for name in sheet_name:
    wb.create_sheet(name)
wb.save('W10_auto.xlsx')

for case in full_list.iter_rows(max_col=14, values_only=True):
    if not case[0] == None and not case[0] == 'Original GM TC ID':
        case_id = case[0]
        test_step = case[7]
        user = case[11]
        connection = case[12]
        sign_status = case[13]
        row = [case_id, test_step]
        if user == 'Driver':
            if connection == 'Online' and sign_status == 'sign_in':
                wb['Online_In'].append(row)
            elif connection == 'Offline' and sign_status == 'sign_in':
                wb['Offline_In'].append(row)
            elif connection == 'Online' and sign_status == 'sign_out':
                wb['Online_Out'].append(row)
            elif connection == 'Offline' and sign_status == 'sign_out':
                wb['Offline_Out'].append(row)
            wb.save('W10_auto.xlsx')